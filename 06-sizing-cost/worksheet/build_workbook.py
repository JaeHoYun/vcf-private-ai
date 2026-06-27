"""
VCF Private AI 사이징·TCO 계산 워크북 생성기.
입력(파란 셀)을 바꾸면 GPU·노드·스토리지·TCO 골격이 수식으로 재계산된다.
출처: https://github.com/JaeHoYun/vcf-private-ai/tree/main/06-sizing-cost
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ARIAL = "Arial"
BLUE = "0000FF"     # 입력(사용자가 바꾸는 값)
BLACK = "000000"    # 수식·계산
GREEN = "008000"    # 다른 시트 링크
HDR_FILL = PatternFill("solid", fgColor="1F3864")
SEC_FILL = PatternFill("solid", fgColor="D9E1F2")
YEL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def hdr(cell, text):
    cell.value = text
    cell.font = Font(name=ARIAL, bold=True, color="FFFFFF", size=12)
    cell.fill = HDR_FILL
    cell.alignment = Alignment(vertical="center")

def section(ws, row, text, span=5):
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=ARIAL, bold=True, color="1F3864")
    c.fill = SEC_FILL
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = SEC_FILL

wb = Workbook()

# ---------- 표지 ----------
cover = wb.active
cover.title = "표지"
cover.sheet_view.showGridLines = False
cover.column_dimensions["A"].width = 2
cover.column_dimensions["B"].width = 100
hdr(cover.cell(row=2, column=2), "VCF Private AI 사이징·TCO 계산 워크북")
notes = [
    "",
    "목적: 입력값만 채우면 GPU·노드·클러스터·스토리지·TCO 골격이 수식으로 자동 산출됩니다.",
    "",
    "[중요 — 반드시 읽으세요]",
    "· 처리량·동시성 등 핵심 입력의 기본값은 '예산 추정 전용 1차 가정치'(부록 A1)이며, 구매 전 윤곽 산정용입니다.",
    "· 확정 사이징은 PoC·부하시험 실측값으로 갈음해야 합니다(가이드 06.5).",
    "· 단가(가격)는 제시하지 않습니다. TCO 시트의 단가 칸은 견적으로 채우세요(부록 A3).",
    "",
    "[색상 약속]",
    "· 파란 글씨 = 입력(여기를 바꿔 시나리오 조정)",
    "· 검은 글씨 = 수식·계산(직접 수정 금지)",
    "· 노란 배경 = 확인 필요 / 1차 가정치 / 견적 입력 칸",
    "",
    "[사용법]",
    "1) '입력' 시트의 파란 셀을 자사 값으로 변경  2) '산정'·'TCO' 시트에서 결과 확인",
    "3) 파일을 Excel/LibreOffice로 열면 수식이 자동 재계산됩니다.",
    "",
    "[적용 범위] 단일 모델을 (Replica당 GPU)×GPU 메모리에 적재하는 추론 워크로드용 1차 도구입니다.",
    "· 모델이 단일 Replica에 안 들어가면 '산정' 시트가 '적재불가'로 표시됩니다 → 입력의 'Replica당 GPU'를 늘리세요.",
    "· 다중 모델 동시 서빙·학습/파인튜닝·MIG 세부·멀티테넌트 합산은 범위 밖입니다.",
    "",
    "[기본 예시] 사내 임직원 5,000명 한국어 문서 RAG 챗봇 → 산출: GPU 약 3장 / 9 노드 / 가용 약 0.6TB",
    "(가이드 docs/08 레퍼런스 시나리오와 동일)",
    "",
    "출처: https://github.com/JaeHoYun/vcf-private-ai/tree/main/06-sizing-cost  ·  라이선스 CC BY 4.0",
    "비공식 문서 — 벤더 공식 입장 아님. 모든 수치는 어림이며 실측·견적으로 재확인 필요.",
]
r = 4
for line in notes:
    c = cover.cell(row=r, column=2, value=line)
    bold = line.startswith("[") or line.startswith("목적")
    c.font = Font(name=ARIAL, bold=bold, size=11, color="C00000" if "중요" in line else BLACK)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------- 입력 ----------
inp = wb.create_sheet("입력")
inp.sheet_view.showGridLines = False
for col, w in zip("ABCDE", (4, 34, 16, 12, 46)):
    inp.column_dimensions[col].width = w
hdr(inp.cell(row=1, column=2), "입력값 (파란 셀을 변경)")
for col in (3, 4, 5):
    inp.cell(row=1, column=col).fill = HDR_FILL

R = {}
row = 2
def add_input(key, label, value, unit="", note="", highlight=False, pct=False, dec=None):
    global row
    inp.cell(row=row, column=2, value=label).font = Font(name=ARIAL)
    c = inp.cell(row=row, column=3, value=value)
    c.font = Font(name=ARIAL, color=BLUE, bold=True)
    c.border = BORDER
    if highlight:
        c.fill = YEL
    if pct:
        c.number_format = "0%"
    elif dec is not None:
        c.number_format = dec
    inp.cell(row=row, column=4, value=unit).font = Font(name=ARIAL, italic=True, color="808080")
    inp.cell(row=row, column=5, value=note).font = Font(name=ARIAL, italic=True, color="808080")
    R[key] = row
    row += 1

def add_section(text):
    global row
    section(inp, row, text)
    row += 1

add_section("워크로드")
add_input("N_users", "대상 사용자 수", 5000, "명")
add_input("dau", "DAU 비율", 0.30, "", "사내 도구 0.2~0.4 (A2)", pct=True)
add_input("q_day", "1인당 일 질의", 20, "회", "A2 기본 5~20")
add_input("hours", "업무시간", 8, "시간")
add_input("peak_ratio", "피크/평균비", 3, "", "A2 기본 3")
add_input("req_time", "요청당 처리시간", 5, "초", "RAG 3~6 (A2)")
add_input("burst", "버스트 헤드룸", 0.30, "", "A2", pct=True)

add_section("모델")
add_input("params", "파라미터 수", 8, "B(십억)")
add_input("prec_w", "정밀도 바이트/param", 2, "byte", "FP16=2 (02.2)")
add_input("layers", "레이어 수", 32, "", "config.json (A2.5)")
add_input("kv_heads", "KV 헤드 수", 8, "", "GQA면 어텐션헤드보다 작음 (A2.5)")
add_input("head_dim", "head_dim", 128, "", "A2.5")
add_input("prec_kv", "KV 정밀도 바이트", 2, "byte", "FP8 KV면 1")

add_section("컨텍스트")
add_input("ctx", "컨텍스트 토큰(입력+출력)", 4096, "tok")
add_input("out_tok", "평균 출력 토큰", 500, "tok")

add_section("GPU·노드 사양")
add_input("gpu_mem", "GPU 메모리", 80, "GB")
add_input("gpu_util", "gpu_memory_utilization", 0.9, "", "vLLM 기본 0.9 (02.1)", pct=True)
add_input("overhead", "활성화·오버헤드", 4, "GB", "어림")
add_input("tput", "[A1] 1차 처리량 가정(단일 GPU)", 2000, "tok/s", "예산 추정 전용·실측 갈음 (A1.1)", highlight=True)
add_input("gpu_per_rep", "Replica당 GPU", 1, "", "텐서 병렬 GPU 수. 가용 VRAM=이 값×GPU메모리. 8B=1, 70B FP16=4 등")
add_input("na_plus", "가용성 추가 Replica(N+1)", 1, "")
add_input("embed_gpu", "임베딩·리랭커 GPU", 1, "", "A1.2 / 03.5")
add_input("gpu_per_node", "노드당 GPU", 1, "", "many-small (04.3)")
add_input("node_head", "노드 헤드룸(오토스케일)", 0.20, "", "04.4", pct=True)
add_input("paif_min", "PAIF 최소 GPU 호스트", 3, "", "01.1 / 03.2")
add_input("gen_workers", "일반 워커 노드", 3, "", "HA")
add_input("cp_nodes", "컨트롤 플레인 노드", 3, "", "홀수 HA (04.1)")
add_input("clusters", "클러스터 수", 1, "", "격리 요구 시 증가 (04.5)")
add_input("sup_cl_lim", "Supervisor 클러스터 한도", 500, "", "VKS 3.6 (04.5)")
add_input("sup_node_lim", "Supervisor 노드 한도", 4000, "", "VKS 3.6 (04.5)")

add_section("스토리지")
add_input("ver_keep", "모델 보존 버전 수", 3, "", "05.1")
add_input("nim_img", "NIM 컨테이너 이미지", 50, "GB", "확인 필요", highlight=True)
add_input("corpus", "문서 코퍼스 원문", 100, "GB")
add_input("corpus_mult", "전처리 배수", 1.5, "", "05.1")
add_input("vec_count", "벡터 수", 5000000, "건")
add_input("vec_dim", "벡터 차원", 1536, "", "05.2")
add_input("vec_bytes", "벡터 float 바이트", 4, "byte", "pgvector float32")
add_input("hnsw_mult", "HNSW 배수", 2, "", "1.5~3 (05.2)")
add_input("logs", "로그·관측(90일)", 30, "GB", "보존 의존")
add_input("prot", "vSAN 보호 오버헤드 배수", 1.5, "", "Auto-RAID (05.3)")
add_input("dedup", "압축·중복제거 절감 배수", 1.0, "", "보수적 1.0 (05.3)")

add_section("라이선스")
add_input("lic_hosts", "라이선스 대상 물리 호스트 수", 3, "", "확인 필요 — 도메인 호스트 수 (A3)", highlight=True)
add_input("cores_host", "호스트당 물리 코어", 64, "", "예시")
add_input("sockets", "호스트당 소켓 수", 2, "", "예시")
add_input("core_min", "소켓당 코어 최소수량", 16, "", "확인 필요 — A3 견적", highlight=True)

def IN(key):
    return f"'입력'!C{R[key]}"

# ---------- 산정 ----------
calc = wb.create_sheet("산정")
calc.sheet_view.showGridLines = False
for col, w in zip("ABCDE", (4, 38, 18, 10, 46)):
    calc.column_dimensions[col].width = w
hdr(calc.cell(row=1, column=2), "사이징 산정 (검은 셀 = 수식, 수정 금지)")
for col in (3, 4, 5):
    calc.cell(row=1, column=col).fill = HDR_FILL

S = {}
crow = 2
def add_calc(key, label, formula, unit="", note="", numfmt="#,##0.0", emphasize=False):
    global crow
    calc.cell(row=crow, column=2, value=label).font = Font(name=ARIAL, bold=emphasize)
    c = calc.cell(row=crow, column=3, value=formula)
    c.font = Font(name=ARIAL, color=BLACK, bold=emphasize)
    c.border = BORDER
    c.number_format = numfmt
    if emphasize:
        c.fill = PatternFill("solid", fgColor="E2EFDA")
    calc.cell(row=crow, column=4, value=unit).font = Font(name=ARIAL, italic=True, color="808080")
    calc.cell(row=crow, column=5, value=note).font = Font(name=ARIAL, italic=True, color="808080")
    S[key] = crow
    crow += 1

def add_calc_section(text):
    global crow
    section(calc, crow, text)
    crow += 1

def C(key):
    return f"C{S[key]}"

add_calc_section("워크로드 → 동시성 (가이드 A2.1)")
add_calc("dau_v", "일 활성 사용자(DAU)", f"={IN('N_users')}*{IN('dau')}", "명", numfmt="#,##0")
add_calc("req_day", "일 요청 수", f"={C('dau_v')}*{IN('q_day')}", "건/일", numfmt="#,##0")
add_calc("peakqps", "피크 QPS", f"={C('req_day')}/({IN('hours')}*3600)*{IN('peak_ratio')}", "req/s", numfmt="#,##0.00")
add_calc("conc_raw", "동시 요청(평균 피크)", f"={C('peakqps')}*{IN('req_time')}", "", numfmt="#,##0.0")
add_calc("conc_design", "설계 피크 동시성", f"=ROUNDUP({C('conc_raw')}*(1+{IN('burst')}),0)", "", numfmt="#,##0", emphasize=True)

add_calc_section("GPU 메모리 (가이드 02)")
add_calc("w_gb", "가중치 메모리", f"={IN('params')}*{IN('prec_w')}", "GB")
add_calc("kv_tok", "토큰당 KV", f"=2*{IN('layers')}*{IN('kv_heads')}*{IN('head_dim')}*{IN('prec_kv')}/1073741824", "GB/tok", numfmt="0.000000")
add_calc("kv_req", "요청당 KV", f"={C('kv_tok')}*{IN('ctx')}", "GB")
add_calc("kv_total", "총 KV(설계 동시성)", f"={C('kv_req')}*{C('conc_design')}", "GB")
add_calc("vram", "소요 VRAM(1 Replica가 전체 동시성 수용 시)", f"={C('w_gb')}+{C('kv_total')}+{IN('overhead')}", "GB", emphasize=True)
add_calc("avail", "가용 VRAM/Replica(gpr×GPU)", f"={IN('gpu_per_rep')}*{IN('gpu_mem')}*{IN('gpu_util')}", "GB")
add_calc("fit", "모델 적재 가능?(gpr 반영)", f'=IF({C("w_gb")}+{IN("overhead")}<={C("avail")},"예","아니오 → Replica당 GPU 늘리기/양자화")', "", numfmt="General")
add_calc("conc_per_gpu", "Replica당 수용 동시성(메모리)", f"=IF({C('kv_req')}<=0,0,FLOOR(MAX({C('avail')}-{C('w_gb')}-{IN('overhead')},0)/{C('kv_req')},1))", "", numfmt="#,##0")
add_calc("fit_warn", "적재 점검", f'=IF({C("fit")}="예",IF({C("conc_per_gpu")}<=0,"KV 여유 없음 → Replica당 GPU 늘리기","정상"),"모델 미적재 → Replica당 GPU 늘리기")', "", numfmt="General")

add_calc_section("처리량 → Replica → GPU (가이드 02.5 / A1)")
add_calc("need_tps", "필요 출력 처리량(피크)", f"={C('peakqps')}*{IN('out_tok')}", "tok/s", numfmt="#,##0")
add_calc("rep_tput", "Replica(처리량 기준)", f"=ROUNDUP({C('need_tps')}/{IN('tput')},0)", "", numfmt="#,##0")
add_calc("rep_mem", "Replica(메모리 기준)", f'=IF({C("conc_per_gpu")}<=0,"적재불가",ROUNDUP({C("conc_design")}/{C("conc_per_gpu")},0))', "", numfmt="#,##0")
add_calc("rep_base", "기본 Replica", f'=IF({C("conc_per_gpu")}<=0,"적재불가",MAX({C("rep_tput")},{C("rep_mem")},1))', "", numfmt="#,##0")
add_calc("rep_total", "총 추론 Replica(+N+1)", f'=IF(ISNUMBER({C("rep_base")}),{C("rep_base")}+{IN("na_plus")},"적재불가")', "", numfmt="#,##0")
add_calc("gpu_infer", "추론 GPU", f'=IF(ISNUMBER({C("rep_total")}),{C("rep_total")}*{IN("gpu_per_rep")},"적재불가")', "", numfmt="#,##0")
add_calc("gpu_total", "총 GPU(논리)", f'=IF(ISNUMBER({C("gpu_infer")}),{C("gpu_infer")}+{IN("embed_gpu")},"적재불가")', "", numfmt="#,##0", emphasize=True)

add_calc_section("노드 → 클러스터 (가이드 04)")
add_calc("gpu_nodes_wl", "GPU 워커 노드(워크로드)", f'=IF(ISNUMBER({C("gpu_total")}),ROUNDUP({C("gpu_total")}/{IN("gpu_per_node")},0),"적재불가")', "", numfmt="#,##0")
add_calc("gpu_nodes", "GPU 워커 노드(PAIF 최소 반영)", f'=IF(ISNUMBER({C("gpu_nodes_wl")}),MAX({C("gpu_nodes_wl")},{IN("paif_min")}),"적재불가")', "", numfmt="#,##0", emphasize=True)
add_calc("gpu_nodes_as", "GPU 워커 노드(오토스케일 상한)", f'=IF(ISNUMBER({C("gpu_nodes")}),ROUNDUP({C("gpu_nodes")}*(1+{IN("node_head")}),0),"적재불가")', "", numfmt="#,##0")
add_calc("cluster_nodes", "클러스터 노드 합계", f'=IF(ISNUMBER({C("gpu_nodes")}),{C("gpu_nodes")}+{IN("gen_workers")}+{IN("cp_nodes")},"적재불가")', "", numfmt="#,##0", emphasize=True)
add_calc("total_nodes", "총 노드(전 클러스터)", f'=IF(ISNUMBER({C("cluster_nodes")}),{C("cluster_nodes")}*{IN("clusters")},"적재불가")', "", numfmt="#,##0")
add_calc("chk_cl", "Supervisor 클러스터 한도", f'=IF({IN("clusters")}<={IN("sup_cl_lim")},"여유","초과")', "", numfmt="General")
add_calc("chk_node", "Supervisor 노드 한도", f'=IF(ISNUMBER({C("total_nodes")}),IF({C("total_nodes")}<={IN("sup_node_lim")},"여유","초과"),"적재불가")', "", numfmt="General")

add_calc_section("스토리지 (가이드 05)")
add_calc("st_model", "모델 아티팩트", f"={C('w_gb')}*{IN('ver_keep')}", "GB")
add_calc("st_corpus", "문서 코퍼스", f"={IN('corpus')}*{IN('corpus_mult')}", "GB")
add_calc("st_vec_raw", "벡터 원시", f"={IN('vec_count')}*{IN('vec_dim')}*{IN('vec_bytes')}/1073741824", "GB")
add_calc("st_vec_hnsw", "벡터 HNSW", f"={C('st_vec_raw')}*{IN('hnsw_mult')}", "GB")
add_calc("st_vec", "벡터 인덱스 합", f"={C('st_vec_raw')}+{C('st_vec_hnsw')}", "GB")
add_calc("st_logical", "논리 합계", f"={C('st_model')}+{IN('nim_img')}+{C('st_corpus')}+{C('st_vec')}+{IN('logs')}", "GB", emphasize=True)
add_calc("st_avail", "필요 가용 용량", f"={C('st_logical')}*{IN('prot')}/{IN('dedup')}", "GB", emphasize=True)

# ---------- TCO ----------
tco = wb.create_sheet("TCO")
tco.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", (4, 30, 16, 18, 22, 30)):
    tco.column_dimensions[col].width = w
hdr(tco.cell(row=1, column=2), "TCO 골격 (단가는 견적으로 — 부록 A3)")
for col in (3, 4, 5, 6):
    tco.cell(row=1, column=col).fill = HDR_FILL

trow = 2
def t_calc(label, formula, unit="", note="", numfmt="#,##0", color=BLACK):
    global trow
    tco.cell(row=trow, column=2, value=label).font = Font(name=ARIAL)
    c = tco.cell(row=trow, column=3, value=formula)
    c.font = Font(name=ARIAL, color=color)
    c.number_format = numfmt
    c.border = BORDER
    tco.cell(row=trow, column=4, value=unit).font = Font(name=ARIAL, italic=True, color="808080")
    tco.cell(row=trow, column=5, value=note).font = Font(name=ARIAL, italic=True, color="808080")
    row_ = trow
    trow += 1
    return row_

section(tco, trow, "라이선스 산정 입력")
trow += 1
r_coreph = t_calc("코어/호스트(최소수량 반영)", f"=MAX({IN('cores_host')},{IN('sockets')}*{IN('core_min')})", "코어", "MAX(실제, 소켓×최소수량)")
r_cores = t_calc("총 라이선스 코어", f"={IN('lic_hosts')}*C{r_coreph}", "코어", "코어 최소수량 반영 (07.2 / A3.1)")
r_gpu = t_calc("NVAIE 대상 GPU", f"=IF(ISNUMBER('산정'!C{S['gpu_total']}),'산정'!C{S['gpu_total']},\"적재불가\")", "GPU", "설치 물리 GPU 전수 (07.2)", color=GREEN)
r_stor = t_calc("스토리지 필요 가용", "='산정'!C" + str(S['st_avail']), "GB", "05.5", color=GREEN)

trow += 1
section(tco, trow, "비용 항목 (단가 칸을 견적으로 채우면 연 환산 자동)")
trow += 1
# 헤더 행
for col, txt in zip((2, 3, 4, 5), ("항목", "수량", "단가(견적 입력)", "연 환산 소계")):
    cc = tco.cell(row=trow, column=col, value=txt)
    cc.font = Font(name=ARIAL, bold=True)
    cc.fill = SEC_FILL
trow += 1

def cost_row(label, qty_formula, note=""):
    global trow
    tco.cell(row=trow, column=2, value=label).font = Font(name=ARIAL)
    q = tco.cell(row=trow, column=3, value=qty_formula)
    q.font = Font(name=ARIAL, color=BLACK)
    q.number_format = "#,##0"
    q.border = BORDER
    p = tco.cell(row=trow, column=4)  # 단가 - 견적 입력
    p.fill = YEL
    p.border = BORDER
    p.number_format = "#,##0"
    sub = tco.cell(row=trow, column=5, value=f'=IF(D{trow}="","견적 입력 필요",IF(ISNUMBER(C{trow}),C{trow}*D{trow},"적재불가"))')
    sub.font = Font(name=ARIAL)
    sub.number_format = "#,##0"
    sub.border = BORDER
    tco.cell(row=trow, column=6, value=note).font = Font(name=ARIAL, italic=True, color="808080")
    trow += 1

cost_row("VCF 코어 구독", f"=C{r_cores}", "코어/년, PAIF 포함(이중계상 금지)")
cost_row("NVAIE(GPU당)", f"=C{r_gpu}", "GPU/년, NVIDIA 별도")
cost_row("DSM", "=1", "VCF entitlement (조건 A3.1)")
cost_row("GPU 하드웨어", f"=C{r_gpu}", "÷ 상각연수 (07.3)")
cost_row("GPU 서버(노드)", f"=IF(ISNUMBER('산정'!C{S['cluster_nodes']}),'산정'!C{S['cluster_nodes']},\"적재불가\")", "÷ 상각연수")
cost_row("스토리지", f"=C{r_stor}", "GB, vSAN 등")
cost_row("네트워크/스위치", "=1", "토폴로지 (05.4)")
cost_row("전력(연)", f"=C{r_gpu}", "GPU 전력×PUE (07.5)")
cost_row("유지보수/지원", "=1", "계약")
cost_row("도입/마이그레이션", "=1", "일회성")

note_c = tco.cell(row=trow + 1, column=2,
    value="단가(노란 칸)는 본 워크북이 제시하지 않습니다. 부록 A3 견적 요청 체크리스트로 채우고 출처를 기록하세요.")
note_c.font = Font(name=ARIAL, italic=True, color="C00000")

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sizing-workbook.xlsx")
wb.save(out)
print("SAVED:", out)
